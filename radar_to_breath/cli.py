from __future__ import annotations

import argparse
import csv
import json
import logging
import os
from pathlib import Path
import sys
import tempfile
import traceback

from openpyxl import load_workbook

from .config import ProcessingConfig
from .edf_io import discover_edf_files
from .pipeline import process_edf


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python -m radar_to_breath",
        description="Convert SleepBRL S1-S16 radar I/Q EDF signals into a 4 Hz candidate breath waveform.",
    )
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--input-dir", type=Path, help="directory searched recursively for sbj*.edf")
    source.add_argument("--edf", type=Path, help="single EDF file")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--subjects-xlsx", type=Path, required=True)
    parser.add_argument("--split", choices=["train", "val", "test"], required=True)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--diagnostics", action="store_true", help="write one diagnostic PNG per record")
    parser.add_argument("--diagnostic-dir", type=Path)
    parser.add_argument("--config", type=Path, help="optional JSON file overriding ProcessingConfig defaults")
    parser.add_argument("--max-records", type=int, default=None, help="process only the first N discovered EDFs")
    parser.add_argument("--log-level", choices=["DEBUG", "INFO", "WARNING", "ERROR"], default="INFO")
    return parser


def _write_reports(output_dir: Path, summaries: list[dict], failures: list[dict]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "run_summary.json").write_text(
        json.dumps(summaries, indent=2, sort_keys=True), encoding="utf-8"
    )
    (output_dir / "failures.json").write_text(
        json.dumps(failures, indent=2, sort_keys=True), encoding="utf-8"
    )
    columns = [
        "record",
        "status",
        "duration_sec",
        "duration_hours",
        "top_iq_pair",
        "top_quality_score",
        "artifact_fraction",
        "bandpower_ratio_0p1_0p35",
        "respiratory_rate_valid_minutes",
        "respiratory_rate_total_minutes",
        "respiratory_rate_median_bpm",
        "respiratory_rate_q1_bpm",
        "respiratory_rate_q3_bpm",
        "selected_iq_pairs",
        "quality_scores",
        "global_fusion_weights",
        "warnings",
        "output_npz",
        "diagnostic_png",
    ]
    with (output_dir / "run_summary.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in summaries:
            serialized = dict(row)
            for key in ("selected_iq_pairs", "quality_scores", "global_fusion_weights", "warnings"):
                serialized[key] = json.dumps(serialized.get(key, []), ensure_ascii=False)
            writer.writerow(serialized)


def _load_subjects(path: Path) -> dict[str, tuple[int, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        positions = {str(value).strip(): index for index, value in enumerate(headers)}
        required = {"Subject", "Sex", "Age"}
        missing = sorted(required - positions.keys())
        if missing:
            raise ValueError(f"subjects workbook is missing columns: {missing}")

        subjects: dict[str, tuple[int, int]] = {}
        for row in rows:
            subject_value = row[positions["Subject"]]
            if subject_value is None:
                continue
            subject = str(subject_value).strip()
            if subject in subjects:
                raise ValueError(f"duplicate subject in workbook: {subject}")

            age_value = float(row[positions["Age"]])
            age = int(age_value)
            if age_value != age:
                raise ValueError(f"non-integer age for {subject}: {age_value}")

            sex_value = str(row[positions["Sex"]]).strip().upper()
            try:
                sex = {"F": 0, "M": 1}[sex_value]
            except KeyError as exc:
                raise ValueError(f"unsupported sex for {subject}: {sex_value!r}") from exc
            subjects[subject] = (age, sex)
        return subjects
    finally:
        workbook.close()


def _write_index(
    output_dir: Path,
    summaries: list[dict],
    subjects: dict[str, tuple[int, int]],
    split: str,
) -> None:
    columns = [
        "path",
        "dataset",
        "source",
        "subject_id",
        "session_id",
        "duration",
        "age",
        "sex",
        "breath_mask",
        "stage_mask",
        "split",
    ]
    with tempfile.NamedTemporaryFile(
        mode="w",
        dir=output_dir,
        prefix=".index_",
        suffix=".csv",
        newline="",
        encoding="utf-8",
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for summary in summaries:
            subject = str(summary["record"])
            age, sex = subjects[subject]
            writer.writerow(
                {
                    "path": summary["output_npz"],
                    "dataset": "sleepbrl",
                    "source": "sleepbrl",
                    "subject_id": subject,
                    "session_id": subject,
                    "duration": 30 * int(summary["stage_epochs"]),
                    "age": age,
                    "sex": sex,
                    "breath_mask": 1,
                    "stage_mask": 1,
                    "split": split,
                }
            )
    try:
        os.replace(temp_path, output_dir / "index.csv")
    finally:
        if temp_path.exists():
            temp_path.unlink()


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    index_path = args.output_dir / "index.csv"
    if index_path.exists():
        if not args.overwrite:
            raise FileExistsError(f"output exists; use --overwrite: {index_path}")
        index_path.unlink()
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(args.output_dir / "run.log", encoding="utf-8"),
        ],
    )
    logger = logging.getLogger("radar_to_breath")
    config = ProcessingConfig.from_json_file(args.config) if args.config else ProcessingConfig()
    files = discover_edf_files(input_dir=args.input_dir, single_edf=args.edf)
    if args.max_records is not None:
        if args.max_records <= 0:
            raise ValueError("--max-records must be positive")
        files = files[: args.max_records]
    subjects = _load_subjects(args.subjects_xlsx)
    discovered_subjects = {path.stem for path in files}
    missing_subjects = sorted(discovered_subjects - subjects.keys())
    if missing_subjects:
        raise ValueError(f"subjects workbook has no rows for records: {missing_subjects}")
    if args.input_dir is not None and args.max_records is None:
        extra_subjects = sorted(subjects.keys() - discovered_subjects)
        if extra_subjects:
            raise ValueError(f"subjects workbook has rows without EDF records: {extra_subjects}")
    logger.info("discovered %d EDF file(s)", len(files))

    summaries: list[dict] = []
    failures: list[dict] = []
    for index, path in enumerate(files, start=1):
        logger.info("[%d/%d] starting %s", index, len(files), path.name)
        try:
            summary = process_edf(
                edf_path=path,
                output_dir=args.output_dir,
                config=config,
                overwrite=args.overwrite,
                diagnostics=args.diagnostics,
                diagnostic_dir=args.diagnostic_dir,
            )
            summaries.append(summary)
            logger.info(
                "[%d/%d] completed %s: top pair=%s, median rate=%.2f bpm",
                index,
                len(files),
                path.name,
                summary["top_iq_pair"],
                summary["respiratory_rate_median_bpm"],
            )
        except Exception as exc:  # failures are recorded and processing continues
            failure = {
                "record": path.stem,
                "path": str(path),
                "error_type": type(exc).__name__,
                "error": str(exc),
                "traceback": traceback.format_exc(),
            }
            failures.append(failure)
            logger.exception("[%d/%d] failed %s", index, len(files), path.name)
        finally:
            _write_reports(args.output_dir, summaries, failures)

    logger.info("finished: %d successful, %d failed", len(summaries), len(failures))
    if not failures:
        _write_index(args.output_dir, summaries, subjects, args.split)
    return 1 if failures else 0
